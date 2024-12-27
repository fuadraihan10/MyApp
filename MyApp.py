from tkinter import *
from tkinter import messagebox,simpledialog
from PIL import Image, ImageTk
import os
import smtplib
import openpyxl
from email.mime.text import MIMEText
import random
import re
import sys
import os



if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

os.chdir(application_path)



images = {}
def load_images():
  try:
    images['google'] = ImageTk.PhotoImage(Image.open(r'Assets/google.png').resize((50, 50)))
    images['facebook'] = ImageTk.PhotoImage(Image.open(r'Assets/facebook.png').resize((41, 41)))
    images['closed_eye_image'] = ImageTk.PhotoImage(Image.open(r'Assets/closed_eye.png').resize((20, 20)))
    images['open_eye_image'] = ImageTk.PhotoImage(Image.open(r'Assets/open_eye.png').resize((20, 20)))
  except Exception as e:
     messagebox.showerror('Error',f"Grafic_loading_error , It might have some grafical errors!!!!\nfor that u might face some problems so first fix the grafical errors\n : {e}")



def create_excel_file(file_name):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.append(["Username", "Email", "Password"])
  wb.save(file_name)



def send_verification_email(email, verification_code):
  try:
    msg = MIMEText(f"Your verification code is: {verification_code}")
    msg['Subject'] = 'Email Verification Code'
    msg['From'] = 'MyAppfuad@gmail.com'  
    msg['To'] = email

    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('MyAppfuad@gmail.com', 'kxsc glyi fibb azmg')  
        server.send_message(msg)
  except Exception as e:
     messagebox.showerror('Error',f"Sending_email_error : {e}")



def is_valid_email(email):
  email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
  return re.match(email_regex, email)



def create_app(boo):
  global app
  app = Tk()
  app.geometry("1100x600")
  app.configure(bg="#FFFFFF")
  app.resizable(False, False)
  icon_image = PhotoImage (file=r"Assets\icon.png")
  app.iconphoto(False, icon_image)

  load_images()
    


  gif = Image.open(r'Assets/gif.gif')

  frames = []
  try:
    while True:
        frames.append(ImageTk.PhotoImage(gif.copy()))
        gif.seek(len(frames))
  except Exception as E:
      pass
  GIF = Label(app)
  GIF.place(x=0,
              y=0,
                height=600,
                  width=600)
  def update_frame(index=0):
    frame = frames[index]
    GIF.configure(image=frame)
    global after_id
    after_id = app.after(50, update_frame, (index + 1) % len(frames))
  update_frame()

def login_page(boo):
  
  if boo:
    create_app(boo=boo)
  app.title("Login-MyApp")

  frame = Frame(app,
                  width=430,
                    height=540,
                      bg="white")
  frame.place(x=635,
                y=30)




  Login_label = Label(frame,
                      text="Login MyApp",
                      fg="#003140",
                      bg="white",
                      font=("Gill Sans Ultra Bold", 23, "bold"))
  Login_label.place(x=82, y=40)




  def username_enter(event):
    if username_entry.get() == "Username":
        username_entry.delete(0, END)

  username_entry = Entry(frame,
                          width=27,
                            font=("Aptos", 14),
                              border=0,
                                fg="black")
  username_entry.place(x=40,
                        y=120)
  username_entry.insert(0, "Username")
  username_entry.bind('<FocusIn>', username_enter)




  def password_enter(event):
      if password_entry.get() == "Password":
          password_entry.delete(0, END)

  password_entry = Entry(frame,
                          width=27,
                            font=("Aptos", 14),
                              border=0,
                                fg="black")
  password_entry.place(x=40, y=170)
  password_entry.insert(0, "Password")
  password_entry.bind('<FocusIn>', password_enter)




  frame1 = Frame(frame,
                  width=350,
                  height=2,
                    bg="black")
  frame1.place(x=40,
                y=143)





  frame2 = Frame(frame,
                  width=350,
                    height=2,
                      bg="black")
  frame2.place(x=40,
                y=193)




  def toggle_password():
    if password_entry.cget('show') == '*':
        password_entry.config(show='')
        eye_button.config(image=images['open_eye_image'])
    else:
        eye_button.config(image=images['closed_eye_image'])
        password_entry.config(show='*')


  eye_button = Button(frame,
                        bd=0,
                          cursor="hand2",
                            background="white",
                      image=images['open_eye_image'],
                        command=toggle_password,
                          borderwidth=0)
  eye_button.place(x=372,
                    y=170)




  def create_reset_password_page():
    app.title("Reset Password Window-MyApp")
    

    frame = Frame(app,
                   width=430,
                   height=540,
                     bg="white")
    frame.place(x=635,
                 y=30)
    reset_label = Label(frame,
                         text="Reset Password",
                           fg="#003140",
                             bg="white",
                               font=("Gill Sans Ultra Bold",
                                      23,
                                      "bold"))
    reset_label.place(x=60,
                       y=40)





    def username_enter(event):
      if username_entry.get()  == "Username or Email":
          username_entry.delete(0, END)

    username_entry = Entry(frame,
                            width=27,
                              font=("Aptos",
                                     14),
                                border=0,
                                  fg="black")
    username_entry.place(x=40,
                          y=120)
    username_entry.insert(0, "Username or Email")
    username_entry.bind('<FocusIn>', username_enter)





    def old_password_enter(event):
      if old_password_entry.get() == "Old Password":
          old_password_entry.delete(0, END)

    old_password_entry = Entry(frame,
                                width=27,
                                  font=("Aptos"
                                        , 14),
                                          border=0,
                                            fg="black",
                                              show='*')
    old_password_entry.place(x=40,
                              y=180)
    old_password_entry.insert(0, "Old Password")
    old_password_entry.bind('<FocusIn>', old_password_enter)





    def new_password_enter(event):
      if new_password_entry.get() == "New Password":
          new_password_entry.delete(0, END)

    new_password_entry = Entry(frame, width=27,
                                font=("Aptos",
                                       14),
                                         border=0,
                                           fg="black",
                                             show='*')
    new_password_entry.place(x=40,
                              y=240)
    new_password_entry.insert(0, "New Password")
    new_password_entry.bind('<FocusIn>', new_password_enter)





    frame1 = Frame(frame,
                    width=350,
                      height=2,
                        bg="black")
    frame1.place(x=40,
                  y=143)

    frame2 = Frame(frame,
                    width=350,
                    height=2,
                      bg="black")
    frame2.place(x=40, y=203)

    frame3 = Frame(frame,
                    width=350,
                      height=2,
                        bg="black")
    frame3.place(x=40,
                  y=263)




    def toggle_old_password():
      if old_password_entry.cget('show') == '*':
          old_password_entry.config(show='')
          toggle_button_old.config(image=images['open_eye_image'])
      else:
          old_password_entry.config(show='*')
          toggle_button_old.config(image=images['closed_eye_image'])

    def toggle_new_password():
      if new_password_entry.cget('show') == '*':
          new_password_entry.config(show='')
          toggle_button_new.config(image=images['open_eye_image'])
      else:
          new_password_entry.config(show='*')
          toggle_button_new.config(image=images['closed_eye_image'])


    toggle_button_old = Button(frame, image=images['open_eye_image'], bg="white", borderwidth=0, command=toggle_old_password)
    toggle_button_old.place(x=370, y=180)
    old_password_entry.config(show="")
    toggle_button_new = Button(frame, image=images['open_eye_image'], bg="white", borderwidth=0, command=toggle_new_password)
    toggle_button_new.place(x=370, y=240)
    new_password_entry.config(show="")





    def reset_password():
      username = username_entry.get()
      old_password = old_password_entry.get()
      new_password = new_password_entry.get()

      error_messages = []

      if username == "" or username == "Username or Email":
          error_messages.append("Username or Email")
      if old_password == "" or old_password == "Old Password":
          error_messages.append("Old Password")
      if new_password == "" or new_password == "New Password":
          error_messages.append("New Password")

      if error_messages:
          if len(error_messages) == 1:
              messagebox.showerror("Error", f"Please fill the {error_messages[0]} field.")
          else:
              messagebox.showerror("Error", f"Please fill the following fields: {', '.join(error_messages)}")
          return

      file_name = "user_data.xlsx"
      if not os.path.isfile(file_name):
          messagebox.showerror("Error", "No users registered yet.")
          return

      wb = openpyxl.load_workbook(file_name)
      ws = wb.active

      try:
          for row in ws.iter_rows(min_row=2, values_only=False):

              if (row[0].value == username or row[1].value == username) and row[2].value == old_password:
                  

                  verification_code = random.randint(100000, 999999)


                  if is_valid_email(username):
                      send_verification_email(username, verification_code)
                      user_code = simpledialog.askstring("Verification", f"Enter the verification code sent to your email ({username}):")
                  else:
                      send_verification_email(str(row[1].value), verification_code)
                      user_code = simpledialog.askstring("Verification", f"Enter the verification code sent to the email {row[1].value}):")


                  if user_code and int(user_code) == verification_code:

                      ws.cell(row=row[0].row, column=3).value = new_password
                      wb.save(file_name)
                      messagebox.showinfo("Success", "Password reset successfully!")
                      return
                  else:
                      messagebox.showerror("Error", "Invalid verification code. Reset failed.")
                      return
          

          messagebox.showerror("Error", "Username/Email or Password is incorrect.")
      
      except Exception as e:
          messagebox.showerror("Error", f"Failed to send verification email. {str(e)}")
      
      finally:
          wb.close()



    reset_button = Button(frame,
                           text="Reset Password",
                             bd=0,
                               fg="white",
                                 bg="#003140", 
                                 cursor="hand2",
                        activebackground="#003140",
                          font=("Cooper Black",
                                 21,
                                 "bold"),
                        activeforeground="white", 
                        command=reset_password)
    reset_button.place(x=80,
                        y=300, 
                        height=65,
                        width=270)




    def back_to_register():
      frame.destroy()
      register_frame(app=app)





    back_button = Button(frame,
                          text="Create New one",
                            bd=0,
                              fg="#003140",
                                bg="white",
                                  cursor="hand2",
                        activebackground="white",
                          font=("Aptos",
                                 11,
                                 "underline"),
                        activeforeground="#003140",
                          command=back_to_register)
    back_button.place(x=210,
                       y=430)





    create_account_label = Label(frame,
                                  text="Don't have an account?",
                                    fg="black",
                                    bg="white",
                                      font=("Aptos",
                                             11))
    create_account_label.place(x=40,
                                y=430)







  def reset_password():
    frame.destroy()
    create_reset_password_page()

  forget_button = Button(frame,
                          text="Forget password?",
                            bd=0,
                              bg="white",
                              command = reset_password,
                                cursor="hand2",
                                  activebackground="white",
                                    font=("Aptos",
                                           11))
  forget_button.place(x=278,
                       y=210)





  def validate_login():
    username = username_entry.get()
    password = password_entry.get()
    if (username == "" and password == "") or (username == "Username" and password == "Password") or (username == "" and password == "Password") or (username == "Username" and password == ""):
      messagebox.showerror("Error","All fields are required!!")
      return
    elif username == "" or username == "Username":
      messagebox.showerror("Error", "Please enter a username.")
      return
    elif password == "" or password == "Password":
      messagebox.showerror("Error", "Please enter a password.")
      return


    file_name = "user_data.xlsx"

    if not os.path.isfile(file_name):
      messagebox.showerror("Error", "No users registered yet.")
      return


    wb = openpyxl.load_workbook(file_name)
    ws = wb.active


    for row in ws.iter_rows(min_row=2, values_only=True):  
        if row[0] == username:  
            if row[2] == password:  
                messagebox.showinfo("Success", "Login Successful!")

                return
            else:
                  messagebox.showerror("Error", "Incorrect password.")
                  return
    else:
        messagebox.showerror("Error", "Username not found.")







  Login_button = Button(frame,
                          text="Login",
                            bd=0,
                              fg="white",
                                bg="#003140",
                                  cursor="hand2",
                        activebackground="#003140",
                          font=("Cooper Black", 25, "bold"),
                        activeforeground="white",
                        command=validate_login)
  Login_button.place(x=80,
                      y=280,
                        height=50,
                          width=280)
  





  or_label = Label(frame,
                    text="________________ OR ________________",
                      font=("Open Sans", 11),
                        bg="white")
  or_label.place(x=90,
                  y=350)
  





  google_Button = Button(frame,
                          image=images['google'],
                            bg="white",
                              border=False,
                          activebackground="white",
                            cursor="hand2",
                          command=lambda: messagebox.showinfo("Login with Google",
                                                              "This feature is not yet available"))
  google_Button.place(x=160,
                        y=390)
  





  facebook_Button = Button(frame,
                            image=images['facebook'],
                              bg="white",
                                border=False,
                            activebackground="white",
                              cursor="hand2",
                            command=lambda: messagebox.showinfo("Login with Facebook",
                                                                "This feature is not yet available"))
  facebook_Button.place(x=230,
                          y=395)
  






  signup_label = Label(frame,
                        text="Don't have an account?",
                          font=("Open Sans",10),
                          bg="white")
  signup_label.place(x=150,
                      y=500)
  




  def new_account_button_function():
    frame.destroy()  
    register_frame(app=app)  

  new_account_button = Button(frame,
                                text="Create new one",
                                  bd=0,
                                    fg="#003140",
                              bg="white",
                                cursor="hand2",
                                  activebackground="white",
                              font=("Open Sans", 10, "underline"),
                                activeforeground="#003140",
                              command=new_account_button_function)
  new_account_button.place(x=300,
                            y=498.7)
  



  if boo:
    app.mainloop()



def register_frame(app):
  app.title("Register-MyApp")



  frame = Frame(app,
                  width=430,
                    height=540,
                      bg="white")
  frame.place(x=635,
                y=30)





  Register_label = Label(frame,
                          text="Register MyApp",
                            fg="#003140",
                              bg="white",
                          font=("Gill Sans Ultra Bold", 21, "bold"))
  Register_label.place(x=82,
                        y=30)






  def email_enter(event):
    if email_entry.get() == "Email":
        email_entry.delete(0, END)

  email_entry = Entry(frame,
                        width=27,
                          font=("Aptos", 13),
                            border=0,
                              fg="black")
  email_entry.place(x=40,
                      y=110)
  
  email_entry.insert(0, "Email")
  email_entry.bind('<FocusIn>', email_enter)





  def Username_enter(event):
    if Username_entry.get() == "Username":
        Username_entry.delete(0, END)

  Username_entry = Entry(frame,
                          width=27,
                            font=("Aptos", 12),
                              border=0,
                                fg="black")
  Username_entry.place(x=40,
                        y=170)
  
  Username_entry.insert(0, "Username")
  Username_entry.bind('<FocusIn>', Username_enter)





  def password_enter(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, END)

  password_entry = Entry(frame, 
                          width=27, 
                          font=("Aptos", 12),
                            border=0, 
                            fg="black")
  password_entry.place(x=40,
                        y=230)
  password_entry.insert(0, "Password")
  password_entry.bind('<FocusIn>', password_enter)






  def Conform_password_enter(event):
    if Conform_password_entry.get() == "Conform Password":
        Conform_password_entry.delete(0, END)

  Conform_password_entry = Entry(frame,
                                  width=27,
                                    font=("Aptos", 12),
                                      border=0,
                                        fg="black")
  Conform_password_entry.place(x=40, 
                                y=290)

  Conform_password_entry.insert(0, "Conform Password")
  Conform_password_entry.bind('<FocusIn>', Conform_password_enter)





  frame1 = Frame(frame, 
                  width=350, 
                  height=2, 
                  bg="black")
  frame1.place(x=40,
                y=142)

  frame2 = Frame(frame, 
                  width=350, 
                  height=2, 
                  bg="black")
  frame2.place(x=40, 
                y=192)

  frame3 = Frame(frame,
                  width=350,
                    height=2, 
                    bg="black")
  frame3.place(x=40,
                y=252)

  frame4 = Frame(frame,
                  width=350,
                    height=2,
                      bg="black")
  frame4.place(x=40,
                y=312)





  def toggle_password():
    if password_entry.cget('show') == '*':
        password_entry.config(show='')
        eye_button.config(image=images['open_eye_image'])
        Conform_password_entry.config(show='')
    else:
        eye_button.config(image=images['closed_eye_image'])
        password_entry.config(show='*')
        Conform_password_entry.config(show='*')





  eye_button = Button(frame,
                        bd=0,
                          cursor="hand2",
                            background="white",
                      image=images['open_eye_image'], 
                      command=toggle_password,
                        borderwidth=0)
  eye_button.place(x=370,
                    y=230)
  





  check=IntVar()
  terms_and_conditions = Checkbutton(frame,
                                      text="I agree to the terms & conditions",
                                      font=("Aptos", 10, "bold"),
                                      background="white",
                                      activebackground="white",
                                      border=False,
                                      cursor="hand2",
                                      variable=check)
  terms_and_conditions.place(x=40,
                              y=325)
  




      
  def register_func():
    email = email_entry.get()
    username = Username_entry.get()
    password = password_entry.get()
    confirm_password = Conform_password_entry.get()

    error_messages = []

    if email == "" or email == "Email":
        error_messages.append("Email")
    if username == "" or username == "Username":
        error_messages.append("Username")
    if password == "" or password == "Password":
        error_messages.append("Password")
    if confirm_password == "" or confirm_password == "Conform Password":
        error_messages.append("Confirm Password")

    if error_messages:
        if len(error_messages) == 1:
            messagebox.showerror("Error", f"Please fill the {error_messages[0]} field.")
        else:
            messagebox.showerror("Error", f"Please fill the following fields: {', '.join(error_messages)}")
        return

    elif password_entry.get() != Conform_password_entry.get():
        messagebox.showerror("Error", "Passwords do not match.")
        return 
    elif check.get() == 0:
        messagebox.showerror("Error", "Please agree to the terms & conditions.")
        return 
    elif not is_valid_email(email_entry.get()):
        messagebox.showerror("Error", "Invalid email format. Please enter a valid email address.")
        return 
    else:
        username = Username_entry.get()
        email = email_entry.get()
        password = password_entry.get()

        file_name = "user_data.xlsx"

        if not os.path.isfile(file_name):
            create_excel_file(file_name)

        wb = openpyxl.load_workbook(file_name)
        ws = wb.active


        for row in ws.iter_rows(min_row=2, values_only=True): 
            existing_username, existing_email, _ = row
            if existing_username == username:
                messagebox.showerror("Error", "Username already exists. Please choose a different username.")
                wb.close()
                return
            if existing_email == email:
                messagebox.showerror("Error", "Email already registered. Please use a different email or login.")
                wb.close()
                return
        else:
            verification_code = random.randint(100000, 999999)


            try:
                send_verification_email(email, verification_code)
                user_code = simpledialog.askstring("Verification", f"Enter the verification code that we have sent to your email :{email}:")

                if user_code and int(user_code) == verification_code:
                    ws.append([username, email, password])
                    wb.save(file_name)
                    messagebox.showinfo("Success", "Registration Successful.")
                else:
                    messagebox.showerror("Error", "Invalid verification code. Registration failed.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to send verification email. {str(e)}")
            finally:
                wb.close()
          




  Register_button = Button(frame,
                            text="Register",
                              bd=0, fg="white",
                                bg="#003140", 
                                cursor="hand2",
                            activebackground="#003140",
                              font=("Cooper Black", 25, "bold"),
                            activeforeground="white",
                            command=register_func)
  Register_button.place(x=100,
                          y=380,
                            height=40,
                              width=250)
  




  signup_label = Label(frame, 
                        text="Don't have an account?", 
                        font=("Open Sans", 10),
                          bg="white")
  signup_label.place(x=155,
                      y=500)
  



  def register_page_fun():
    frame.destroy()  
    login_page(boo=False)  
      



  log_in_button = Button(frame, 
                          text="Login", 
                          bd=0, fg="#003140", 
                          bg="white", 
                          cursor="hand2",
                          activebackground="white", 
                          font=("Open Sans", 10, "underline"),
                          activeforeground="#003140", 
                          command=register_page_fun)
  log_in_button.place(x=300, 
                      y=498.7)
    


login_page(boo=True)